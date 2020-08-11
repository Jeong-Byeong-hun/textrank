from newspaper import Article
from konlpy.tag import Kkma
from konlpy.tag import Twitter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.preprocessing import normalize
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
import requests
import pandas as pd
import re
import numpy as np
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db

title_text = []
link_text = []
source_text = []
date_text = []
contents_text = []
result = {}

RESULT_PATH ='C:/Users/wjdqu/Desktop/pyexcel/'  #결과 저장할 경로
now = datetime.now()


def date_cleansing(test):
    try:
        # 지난 뉴스
        # 머니투데이  10면1단  2018.11.05.  네이버뉴스   보내기
        pattern = '\d+.(\d+).(\d+).'  # 정규표현식

        r = re.compile(pattern)
        match = r.search(test).group(0)  # 2018.11.05.
        date_text.append(match)

    except AttributeError:
        # 최근 뉴스
        # 이데일리  1시간 전  네이버뉴스   보내기
        pattern = '\w* (\d\w*)'  # 정규표현식

        r = re.compile(pattern)
        match = r.search(test).group(1)
        # print(match)
        date_text.append(match)

#내용 정제화 함수
def contents_cleansing(contents):
    first_cleansing_contents = re.sub('<dl>.*?</a> </div> </dd> <dd>', '',
                                      str(contents)).strip()  #앞에 필요없는 부분 제거
    second_cleansing_contents = re.sub('<ul class="relation_lst">.*?</dd>', '',
                                       first_cleansing_contents).strip()#뒤에 필요없는 부분 제거 (새끼 기사)
    third_cleansing_contents = re.sub('<.+?>', '', second_cleansing_contents).strip()
    contents_text.append(third_cleansing_contents)
    #print(contents_text)


def crawler(maxpage, query, sort, s_date, e_date):
    s_from = s_date.replace(".", "")
    e_to = e_date.replace(".", "")
    page = 1
    maxpage_t = (int(maxpage) - 1) * 10 + 1  # 11= 2페이지 21=3페이지 31=4페이지  ...81=9페이지 , 91=10페이지, 101=11페이지

    while page <= maxpage_t:
        url = "https://search.naver.com/search.naver?where=news&query=" + query + "&sort=" + sort + "&ds=" + s_date + "&de=" + e_date + "&nso=so%3Ar%2Cp%3Afrom" + s_from + "to" + e_to + "%2Ca%3A&start=" + str(
            page)

        response = requests.get(url)
        html = response.text

        # 뷰티풀소프의 인자값 지정
        soup = BeautifulSoup(html, 'html.parser')

        # <a>태그에서 제목과 링크주소 추출
        atags = soup.select('._sp_each_title')
        for atag in atags:
            title_text.append(atag.text)  # 제목
            link_text.append(atag['href'])  # 링크주소


         # 신문사 추출
        source_lists = soup.select('._sp_each_source')
        for source_list in source_lists:
            source_text.append(source_list.text)  # 신문사

        # 날짜 추출
        date_lists = soup.select('.txt_inline')
        for date_list in date_lists:
            test = date_list.text
            date_cleansing(test)  # 날짜 정제 함수사용

        # 본문요약본
        contents_lists = soup.select('ul.type01 dl')
        for contents_list in contents_lists:
            # print('==='*40)
            # print(contents_list)
            contents_cleansing(contents_list)  # 본문요약 정제화

        # print(len(link_text), len(title_text), len(source_text),len(date_text)) 길이를 찍은 결과 17 10 10 10  link 길이가 17이 나와서 오류가 난 거같음

        # 모든 리스트 딕셔너리형태로 저장
        result = {"date": date_text, "title": title_text, "source": source_text, "contents": contents_text,
                  "link": link_text}
        print(page)

        df = pd.DataFrame(result)  # df로 변환
        page += 10

    # 새로 만들 파일이름 지정
    outputFileName = '%s-%s-%s  %s시 %s분 %s초 merging.xlsx' % (
    now.year, now.month, now.day, now.hour, now.minute, now.second)
    df.to_excel(RESULT_PATH + outputFileName, sheet_name='sheet1')


def main():
    info_main = input("=" * 50 + "\n" + "입력 형식에 맞게 입력해주세요." + "\n" + " 시작하시려면 Enter를 눌러주세요." + "\n" + "=" * 50)

    maxpage = input("최대 크롤링할 페이지 수 입력하시오: ")
    query = input("검색어 입력: ")
    sort = input("뉴스 검색 방식 입력(관련도순=0  최신순=1  오래된순=2): ")  # 관련도순=0  최신순=1  오래된순=2
    s_date = input("시작날짜 입력(2019.04.25):")  # 2019.01.04
    e_date = input("끝날짜 입력(2019.04.26):")  # 2019.01.05

    crawler(maxpage, query, sort, s_date, e_date)

# 뉴스 요약 알고리즘

class SentenceTokenizer(object):

    def __init__(self):

        self.kkma = Kkma()
        self.twitter = Twitter()
        self.stopwords = ['중인', '만큼', '마찬가지', '꼬집었', "연합뉴스", "데일리", "동아일보", "중앙일보", "조선일보", "기자"
            , "아", "휴", "아이구", "아이쿠", "아이고", "어", "나", "우리", "저희", "따라", "의해", "을", "를", "에", "의", "가"
            , "으로", "로", "에게", "뿐이다", "의거하여", "근거하여", "입각하여", "기준으로", "예하면", "예를", "들면", "들자면"
            , "저", "소인", "소생", "저희", "지말고", "하지마", "하지마라", "다른", "물론", "또한", "그리고", "비길수", "없다"
            , "해서는", "안된다", "뿐만", "아니라", "만이", "아니다", "만은", "아니다", "막론하고", "관계없이", "그치지", "않다"
            , "그러나", "그런데", "하지만", "든간에", "논하지", "않다", "따지지", "않다", "설사", "비록", "더라도", "아니면"
            , "만", "못하다", "하는", "편이", "낫다", "불문하고", "향하여", "향해서", "향하다", "쪽으로", "틈타", "이용하여"
            , "타다", "오르다", "제외하고", "이", "외에", "이", "밖에", "하여야", "비로소", "한다면", "몰라도", "외에도"
            , "이곳", "여기", "부터", "기점으로", "따라서", "할", "생각이다", "하려고하다", "이리하여", "그리하여", "그렇게"
            , "함으로써", "하지만", "일때", "할때", "앞에서", "중에서", "보는데서", "으로써", "로써", "까지", "해야한다", "일것이다"
            , "반드시", "할줄알다", "할수있다", "할수있어", "임에", "틀림없다", "한다면", "등", "등등", "제", "겨우", "단지", "다만"
            , "할뿐", "딩동", "댕그", "대해서", "대하여", "대하면", "훨씬", "얼마나", "얼마만큼", "얼마큼", "남짓", "여", "얼마간"
            , "약간", "다소", "좀", "조금", "다수", "몇", "얼마", "지만", "하물며", "또한", "그러나", "그렇지만", "하지만", "이외에도"
            , "대해", "말하자면", "뿐이다", "다음에", "반대로", "반대로", "말하자면", "이와", "반대로", "바꾸어서", "말하면", "바꾸어서"
            , "한다면", "만약", "그렇지않으면", "까악", "툭", "딱", "삐걱거리다", "보드득", "비걱거리다", "꽈당", "응당", "해야한다", "에"
            , "가서", "각", "각각", "여러분", "각종", "각자", "제각기", "하도록하다", "와", "과", "그러므로", "그래서", "고로"
            , "까닭에", "하기", "때문에", "거니와", "이지만", "대하여", "관하여", "관한", "과연", "실로", "아니나다를가", "생각한대로"
            , "진짜로", "한적이있다", "하곤하였다", "하", "하하", "허허", "아하", "거바", "와", "오", "왜", "어째서", "무엇때문에", "어찌"
            , "하겠는가", "무슨", "어디", "어느곳", "더군다나", "하물며", "더욱이는", "어느때", "언제", "야", "이봐", "어이", "여보시오"
            , "흐흐", "흥", "휴", "헉헉", "헐떡헐떡", "영차", "여차", "어기여차", "끙끙", "아야", "앗", "아야", "콸콸", "졸졸", "좍좍"
            , "뚝뚝", "주룩주룩", "솨", "우르르", "그래도", "또", "그리고", "바꾸어말하면", "바꾸어말하자면", "혹은", "혹시", "답다"
            , "및", "그에", "따르는", "때가", "되어", "즉", "지든지", "설령", "가령", "하더라도", "할지라도", "일지라도", "지든지", "몇"
            , "거의", "하마터면", "인젠", "이젠", "된바에야", "된이상", "만큼 어찌됏든", "그위에", "게다가", "점에서", "보아", "비추어"
            , "보아", "고려하면", "하게될것이다", "일것이다", "비교적", "좀", "보다더", "비하면", "시키다", "하게하다", "할만하다", "의해서"
            , "연이서", "이어서", "잇따라", "뒤따라", "뒤이어", "결국", "의지하여", "기대여", "통하여", "자마자", "더욱더", "불구하고", "얼마든지"
            , "마음대로", "주저하지", "않고", "곧", "즉시", "바로", "당장", "하자마자", "밖에", "안된다", "하면된다", "그래", "그렇지", "요컨대", "다시"
            , "말하자면", "바꿔", "말하면", "즉", "구체적으로", "말하자면", "시작하여", "시초에", "이상", "허", "헉", "허걱", "바와같이", "해도좋다", "해도된다"
            , "게다가", "더구나", "하물며", "와르르", "팍", "퍽", "펄렁", "동안", "이래", "하고있었다", "이었다", "에서", "로부터", "까지", "예하면", "했어요"
            , "해요", "함께", "같이", "더불어", "마저", "마저도", "양자", "모두", "습니다", "가까스로", "하려고하다", "즈음하여", "다른", "다른", "방면으로"
            , "해봐요", "습니까", "했어요", "말할것도", "없고", "무릎쓰고", "개의치않고", "하는것만", "못하다", "하는것이", "낫다", "매", "매번", "들", "모"
            , "어느것", "어느", "로써", "갖고말하자면", "어디", "어느쪽", "어느것", "어느해", "어느", "년도", "라", "해도", "언젠가", "어떤것", "어느것"
            , "저기", "저쪽", "저것", "그때", "그럼", "그러면", "요만한걸", "그래", "그때", "저것만큼", "그저", "이르기까지", "할", "줄", "안다", "할", "힘이"
            , "있다", "너", "너희", "당신", "어찌", "설마", "차라리", "할지언정", "할지라도", "할망정", "할지언정", "구토하다", "게우다", "토하다", "메쓰겁다"
            , "옆사람", "퉤", "쳇", "의거하여", "근거하여", "의해", "따라", "힘입어", "그", "다음", "버금", "두번째로", "기타", "첫번째로", "나머지는", "그중에서"
            , "견지에서", "형식으로", "쓰여", "입장에서", "위해서", "단지", "의해되다", "하도록시키다", "뿐만아니라", "반대로", "전후", "전자", "앞의것", "잠시"
            , "잠깐", "하면서", "그렇지만", "다음에", "그러한즉", "그런즉", "남들", "아무거나", "어찌하든지", "같다", "비슷하다", "예컨대", "이럴정도로", "어떻게"
            , "만약", "만일", "위에서", "서술한바와같이", "인", "듯하다", "하지", "않는다면", "만약에", "무엇", "무슨", "어느", "어떤", "아래윗", "조차", "한데"
            , "그럼에도", "불구하고", "여전히", "심지어", "까지도", "조차도", "하지", "않도록", "않기", "위하여", "때", "시각", "무렵", "시간", "동안", "어때"
            , "어떠한", "하여금", "네", "예", "우선", "누구", "누가", "알겠는가", "아무도", "줄은모른다", "줄은", "몰랏다", "하는", "김에", "겸사겸사", "하는바"
            , "그런", "까닭에", "한", "이유는", "그러니", "그러니까", "때문에", "그", "너희", "그들", "너희들", "타인", "것", "것들", "너", "위하여", "공동으로"
            , "동시에", "하기", "위하여", "어찌하여", "무엇때문에", "붕붕", "윙윙", "나", "우리", "엉엉", "휘익", "윙윙", "오호", "아하", "어쨋든", "만"
            , "못하다    하기보다는", "차라리", "하는", "편이", "낫다", "흐흐", "놀라다", "상대적으로", "말하자면", "마치", "아니라면", "쉿", "그렇지", "않으면"
            , "그렇지", "않다면", "안", "그러면", "아니었다면", "하든지", "아니면", "이라면", "좋아", "알았어", "하는것도", "그만이다", "어쩔수", "없다", "하나"
            , "일", "일반적으로", "일단", "한켠으로는", "오자마자", "이렇게되면", "이와같다면", "전부", "한마디", "한항목", "근거로", "하기에", "아울러", "하지"
            , "않도록", "않기", "위해서", "이르기까지", "이", "되다", "로", "인하여", "까닭으로", "이유만으로", "이로", "인하여", "그래서", "이", "때문에",
                          "그러므로"
            , "그런", "까닭에", "알", "수", "있다", "결론을", "낼", "수", "있다", "으로", "인하여", "있다", "어떤것", "관계가", "있다", "관련이", "있다",
                          "연관되다"
            , "어떤것들", "에", "대해", "이리하여", "그리하여", "여부", "하기보다는", "하느니", "하면", "할수록", "운운", "이러이러하다", "하구나", "하도다"
            , "다시말하면", "다음으로", "에", "있다", "에", "달려", "있다", "우리", "우리들", "오히려", "하기는한데", "어떻게", "어떻해", "어찌됏어", "어때"
            , "어째서", "본대로", "자", "이", "이쪽", "여기", "이것", "이번", "이렇게말하자면", "이런", "이러한", "이와", "같은", "요만큼", "요만한", "것",
                          "얼마"
            , "안", "되는", "것", "이만큼", "이", "정도의", "이렇게", "많은", "것", "이와", "같다", "이때", "이렇구나", "것과", "같이", "끼익", "삐걱",
                          "따위"
            , "와", "같은", "사람들", "부류의", "사람들", "왜냐하면", "중의하나", "오직", "오로지", "에", "한하다", "하기만", "하면", "도착하다", "까지", "미치다"
            , "도달하다", "정도에", "이르다", "할", "지경이다", "결과에", "이르다", "관해서는", "여러분", "하고", "있다", "한", "후", "혼자", "자기", "자기집"
            , "자신", "우에", "종합한것과같이", "총적으로", "보면", "총적으로", "말하면", "총적으로", "대로", "하다", "으로서", "참", "그만이다", "할", "따름이다"
            , "쿵", "탕탕", "쾅쾅", "둥둥", "봐", "봐라", "아이야", "아니", "와아", "응", "아이", "참나", "년", "월", "일", "령", "영", "일", "이",
                          "삼", "사"
            , "오", "육", "륙", "칠", "팔", "구", "이천육", "이천칠", "이천팔", "이천구", "하나", "둘", "셋", "넷", "다섯", "여섯", "일곱", "여덟",
                          "아홉", "령", "영", "라며"]

    def url2sentences(self, url):
        article = Article(url, language='ko')
        article.download()
        article.parse()
        sentences = self.kkma.sentences(article.text)
        for idx in range(0, len(sentences)):
            if len(sentences[idx]) <= 10:
                sentences[idx - 1] += (' ' + sentences[idx])
                sentences[idx] = ''

        return sentences

    def text2sentences(self, text):
        sentences = self.kkma.sentences(text)
        for idx in range(0, len(sentences)):
            if len(sentences[idx]) <= 10:
                sentences[idx - 1] += (' ' + sentences[idx])
                sentences[idx] = ''

        return sentences

    def get_nouns(self, sentences):
        nouns = []
        for sentence in sentences:
            if sentence is not '':
                nouns.append(' '.join([noun for noun in self.twitter.nouns(str(sentence))
                                       if noun not in self.stopwords and len(noun) > 1]))
        return nouns


class GraphMatrix(object):
    def __init__(self):
        self.tfidf = TfidfVectorizer()
        self.cnt_vec = CountVectorizer()
        self.graph_sentence = []

    def bulid_sent_graph(self, sentence):
        tfidf_mat = self.tfidf.fit_transform(sentence).toarray()
        self.graph_sentence = np.dot(tfidf_mat, tfidf_mat.T)
        return self.graph_sentence

    def build_words_graph(self, sentence):
        cnt_vec_mat = normalize(self.cnt_vec.fit_transform(sentence).toarray().astype(float), axis=0)
        vocab = self.cnt_vec.vocabulary_
        return np.dot(cnt_vec_mat.T, cnt_vec_mat), {vocab[word]: word for word in vocab}


class Rank(object):

    def get_ranks(self, graph, d=0.85):
        A = graph
        matrix_size = A.shape[0]
        for id in range(matrix_size) :
            A[id, id] = 0
            link_sum = np.sum(A[:, id])
            if link_sum != 0:
                A[:, id] /= link_sum
            A[:, id] *= -d
            A[id, id] = 1
        B = (1-d) * np.ones((matrix_size, 1))
        ranks = np.linalg.solve(A, B)
        return {idx: r[0] for idx, r in enumerate(ranks)}


class TextRank(object):

    def __init__(self, text):
        self.sent_tokenize = SentenceTokenizer()

        if text[:5] in ('http:', 'https'):
            self.sentences = self.sent_tokenize.url2sentences(text)
        else:
            self.sentences = self.sent_tokenize.text2sentences(text)

        self.nouns = self.sent_tokenize.get_nouns(self.sentences)

        self.graph_matrix = GraphMatrix()
        self.sent_graph = self.graph_matrix.bulid_sent_graph(self.nouns)
        self.words_graph, self.idx2word = self.graph_matrix.build_words_graph(self.nouns)

        self.rank = Rank()
        self.sent_rank_idx = self.rank.get_ranks(self.sent_graph)

        self.sorted_sent_rank_idx = sorted(self.sent_rank_idx, key=lambda k: self.sent_rank_idx[k], reverse=True)
        self.word_rank_idx = self.rank.get_ranks(self.words_graph)
        self.sorted_word_rank_idx = sorted(self.word_rank_idx, key=lambda k: self.word_rank_idx[k], reverse=True)

    def summarize(self, sent_num=3):
        summary = []

        index = []
        for idx in self.sorted_sent_rank_idx[:sent_num]:
            index.append(idx)
        index.sort()
        for idx in index:
            summary.append(self.sentences[idx])
        return summary

    def keywords(self, word_num=5):
        rank = Rank()

        rank_idx = rank.get_ranks(self.words_graph)
        sorted_rank_idx = sorted(rank_idx, key=lambda k: rank_idx[k], reverse=True)
        keywords = []
        index = []
        for idx in sorted_rank_idx[:word_num]:
            index.append(idx)
        # index.sort()
        for idx in index:
            keywords.append(self.idx2word[idx])
        return keywords

# main()

# 읽어올 경로 설정 및 읽을 셀 설정
load_wb = load_workbook("C:/Users/wjdqu/Desktop/pyexcel/pycharm12.xlsx", data_only=True)
load_ws = load_wb['sheet1']
url = load_ws.cell(2, 6).value
i = 2
cred = credentials.Certificate('C:/Users/wjdqu/Desktop/GG/pythondata-3ec80-firebase-adminsdk-jek4z-71d665cab3.json')

firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://pythondata-3ec80.firebaseio.com'
})
ref = db.reference()
user_ref = ref.child('data')
while url is not None:
    textrank = TextRank(url)
    sum = []
    for row in textrank.summarize(3):
        sum += row
        print(row)
        print()
        # print('keywords :', textrank.keywords())
    user_ref.push().set({ #데이터베이스에 넣는 과정
        'url' : {
            'link': url,
         },
        'summarize': {
            'textsum': ''.join(sum)
        },
        'word': {
            'keyword': textrank.keywords()
            }}
        )
    i = i + 1
    url = load_ws.cell(i, 6).value
